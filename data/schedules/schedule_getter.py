import requests
from openpyxl import Workbook
import datetime


def get_match_timestamp(match):
    try:
        stamp = int(match["scheduledTime"]["startTime"])
        return stamp
    except:
        return 1633440800000


excel_filename = "Schedules.xlsx"
divisions = {'storm': 'S', 'heroic': 'H', 'nexus': 'N', 'a-east': 'AE', 'a-west': 'AW',
             'b-northeast': 'BNE', 'b-southeast': 'BSE', 'b-west': 'BW', 'c-east': 'CE',
             'c-west': 'CW', 'd-east': 'DE', 'd-west': 'DW', 'e-east': 'EE', 'e-west': 'EW'}

workbook = Workbook()

url = "http://nexusgamingseries.org/api/schedule/fetch/division/matches/"

for div in divisions.keys():
    print("=============", div, "=============")
    r = requests.post(url, data={"season": 11, "division": div})
    worksheet = workbook.create_sheet(divisions[div])
    worksheet.cell(1, 1, "Home")
    worksheet.cell(1, 2, "Away")
    worksheet.cell(1, 3, "Score Home")
    worksheet.cell(1, 4, "Score Away")
    worksheet.cell(1, 5, "Forfeit")
    worksheet.cell(1, 6, "Date")
    matches = r.json()['returnObject']

    matches = sorted(matches, key=lambda x: get_match_timestamp(x))
    row_number = 2
    for match in matches:
        reported = False
        forfeit = 0
        if "forfeit" in match.keys() and match["forfeit"]:
            forfeit = 1
        home_score = 0
        away_score = 0
        if "reported" in match.keys() and match["reported"]:
            try:
                reported = True
                home_score = match["home"]["score"]
                away_score = match["away"]["score"]
            except KeyError:
                print(match["home"]["ticker"], match["away"]["ticker"])
                home_score = 1
                away_score = 2
        home_ticker = ""
        away_ticker = ""
        try:
            home_ticker = match["home"]["ticker"]
        except KeyError:
            home_ticker = match["home"]["teamName"][0:5]
        try:
            away_ticker = match["away"]["ticker"]
        except KeyError:
            away_ticker = match["away"]["teamName"][0:5]
        try:
            stamp = int(match["scheduledTime"]["startTime"])
        except KeyError:
            stamp = 1633440800000

        if str(home_ticker).startswith("="):
            home_ticker = home_ticker.replace("=", "-")
        if str(away_ticker).startswith("="):
            away_ticker = away_ticker.replace("=", "-")
        match_date = datetime.datetime.fromtimestamp(stamp // 1000)
        worksheet.cell(row_number, 1, home_ticker)
        worksheet.cell(row_number, 2, away_ticker)
        if home_score > 0 or away_score > 0:
            worksheet.cell(row_number, 3, home_score)
            worksheet.cell(row_number, 4, away_score)
            worksheet.cell(row_number, 5, forfeit)
            worksheet.cell(row_number, 6, match_date)
        row_number = row_number + 1
workbook.remove_sheet(workbook.get_sheet_by_name("Sheet"))

workbook.save(excel_filename)
workbook.close()
